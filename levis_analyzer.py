import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go
import plotly.express as px

st.set_page_config(page_title="Levi's Sale Analyzer", layout="wide", page_icon="📊")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Plus+Jakarta+Sans:wght@600;700;800&display=swap');

*, *::before, *::after { font-family: 'Inter', sans-serif !important; box-sizing: border-box; }
.stApp { background: #f4f0ff !important; }
#MainMenu, footer, header { visibility: hidden; }
[data-testid="stToolbar"], [data-testid="stDecoration"], [data-testid="stStatusWidget"] { display: none !important; }

.block-container { padding-top: 0.8rem !important; padding-bottom: 1rem !important; }
[data-testid="stAppViewContainer"] > section > div { padding-top: 0 !important; }

/* ══ NAVBAR ══ */
.hero {
    padding: 0.55rem 1.4rem;
    display: flex; align-items: center; gap: 1rem;
    background: linear-gradient(90deg, #3a0068 0%, #6a1b9a 55%, #9c27b0 100%);
    margin-bottom: 1rem; border-radius: 12px;
    box-shadow: 0 3px 14px rgba(106,27,154,0.3);
}
.hero-badge {
    background: rgba(255,255,255,0.18); border: 1.5px solid rgba(255,255,255,0.35);
    color: #ffffff; font-size:.56rem; font-weight:700; letter-spacing:2px;
    text-transform:uppercase; padding:4px 11px; border-radius:20px;
    white-space:nowrap; flex-shrink:0;
}
.hero-divider { width:1px; height:26px; background:rgba(255,255,255,0.22); flex-shrink:0; }
.hero-title {
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-size: 1.05rem; font-weight: 800; color: #ffffff;
    margin: 0; line-height: 1; white-space:nowrap; flex-shrink:0;
}
.hero-arrow { color:rgba(255,255,255,0.45); font-size:.95rem; flex-shrink:0; }
.hero-sub-line {
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-size: .8rem; font-weight: 600; color: #e8c8ff;
    white-space:nowrap; overflow:hidden; text-overflow:ellipsis;
}
.hero-sub { color: rgba(255,255,255,0.52); font-size: .65rem; margin: 0.08rem 0 0 0; font-weight: 400; }

/* ══ KPI CARDS ══ */
.kpi-card {
    background: linear-gradient(135deg, #6a1b9a 0%, #9c27b0 100%);
    border: none; border-radius: 16px; padding: 1.2rem 1.4rem;
    box-shadow: 0 4px 18px rgba(106,27,154,0.35);
    transition: box-shadow 0.2s, transform 0.2s;
}
.kpi-card:hover { box-shadow: 0 8px 28px rgba(106,27,154,0.45); transform: translateY(-3px); }
.kpi-label { font-size:.6rem; font-weight:700; letter-spacing:2.5px; text-transform:uppercase; color:rgba(255,255,255,0.75); margin-bottom:.5rem; }
.kpi-value { font-family:'Plus Jakarta Sans', sans-serif !important; font-size:1.75rem; font-weight:800; color:#ffffff; line-height:1; }
.kpi-sub { font-size:.75rem; color:rgba(255,255,255,0.7); margin-top:.35rem; font-weight:500; }

/* ══ SECTION TITLES ══ */
.section-title {
    font-size:.65rem; font-weight:700; letter-spacing:2.5px; text-transform:uppercase;
    color:#6a1b9a; padding:.5rem 0; margin-bottom:.8rem;
    border-bottom: 2.5px solid #ddd6fe;
}

/* ══ STREAMLIT OVERRIDES ══ */
.stApp > div { background: #f4f0ff !important; }
p { color: #1a0030 !important; font-size:.9rem !important; }
label { color: #3d0066 !important; font-weight:600 !important; font-size:.85rem !important; }
[data-testid="stWidgetLabel"] p { color:#6a1b9a !important; font-size:.85rem !important; font-weight:600 !important; }
div[data-testid="stDataFrame"] * { color:#1a0030 !important; font-size:.84rem !important; }

.stSelectbox > div > div, .stMultiSelect > div > div {
    background:#ffffff !important; border:1.5px solid #c084fc !important; border-radius:10px !important;
    color:#1a0030 !important;
}
[data-baseweb="popover"] { background:#fff !important; border:1px solid #c084fc !important; }
[data-baseweb="popover"] * { color:#1a0030 !important; background:#fff !important; }
[data-baseweb="option"]:hover { background:rgba(106,27,154,0.08) !important; }
li[aria-selected="true"] { background:rgba(106,27,154,0.12) !important; color:#6a1b9a !important; }

[data-baseweb="tag"] { background:#ede9fe !important; }
[data-baseweb="tag"] span { color:#4c1d95 !important; font-weight:600 !important; }

.stButton > button {
    background: linear-gradient(135deg,#6a1b9a,#9c27b0) !important;
    color: #ffffff !important; border: none !important; border-radius: 12px !important;
    font-weight: 700 !important; font-size: .95rem !important; padding: .7rem 2rem !important;
    box-shadow: 0 4px 14px rgba(106,27,154,0.38) !important;
    transition: all 0.2s !important; letter-spacing: .4px !important;
    display: flex !important; align-items: center !important; justify-content: center !important;
    text-shadow: 0 1px 2px rgba(0,0,0,0.2) !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg,#7b1fa2,#ab47bc) !important;
    box-shadow: 0 6px 22px rgba(106,27,154,0.48) !important;
    transform: translateY(-1px) !important; color: #ffffff !important;
}
.stButton > button p { color: #ffffff !important; font-weight: 700 !important; }
.stDownloadButton > button {
    background:#fff !important; color:#6a1b9a !important;
    border:2px solid #6a1b9a !important; border-radius:10px !important;
    font-weight:700 !important; font-size:.88rem !important;
}
.stDownloadButton > button:hover { background:#f5f0ff !important; }

[data-testid="stFileUploader"] { background: #ffffff !important; border: none !important; border-radius: 14px !important; padding: 0 !important; }
[data-testid="stFileUploaderDropzone"] {
    background: #ffffff !important; border: 2px dashed #c084fc !important; border-radius: 14px !important;
    text-align: center !important; padding: 16px !important; display: flex !important;
    flex-direction: column !important; align-items: center !important; justify-content: center !important;
    min-height: 0 !important; max-height: 80px !important;
}
[data-testid="stFileUploaderDropzone"] svg { fill: #9c27b0 !important; }
[data-testid="stFileUploaderDropzone"] > div { display: flex !important; flex-direction: column !important; align-items: center !important; justify-content: center !important; width: 100% !important; }
[data-testid="stFileUploaderDropzone"] button { visibility: hidden !important; height: 0 !important; padding: 0 !important; margin: 0 !important; }
[data-testid="stFileUploadDeleteBtn"] { visibility: visible !important; display: flex !important; align-items: center !important; }
[data-testid="stFileUploadDeleteBtn"] button {
    visibility: visible !important; display: inline-flex !important; align-items: center !important; height: 28px !important;
    background: #fee2e2 !important; border: 1.5px solid #fca5a5 !important; border-radius: 6px !important;
    padding: 0 12px !important; cursor: pointer !important; color: #dc2626 !important; font-size: 12px !important; font-weight: 700 !important;
}
[data-testid="stFileUploadDeleteBtn"] button:hover { background: #fecaca !important; }
[data-testid="stFileUploadDeleteBtn"] svg { fill: #dc2626 !important; }

.stTabs [data-baseweb="tab-list"] { background:#fff !important; border-radius:12px !important; padding:4px !important; border:1.5px solid #ddd6fe !important; box-shadow: 0 2px 8px rgba(106,27,154,0.08) !important; }
.stTabs [data-baseweb="tab"] { color: #1a0030 !important; border-radius:8px !important; font-size:.84rem !important; font-weight:600 !important; }
.stTabs [aria-selected="true"] { background: linear-gradient(135deg,#6a1b9a,#9c27b0) !important; color: #ffffff !important; font-weight:700 !important; }
.stTabs [aria-selected="true"] * { color: #ffffff !important; }

.stSuccess { background:#f0fdf4 !important; border:1px solid #86efac !important; border-radius:10px !important; }
.stSuccess * { color:#166534 !important; font-weight:600 !important; }
[data-testid="stSpinner"] * { color:#6a1b9a !important; }
.stDataFrame { border-radius:12px !important; overflow:hidden; }
</style>
""", unsafe_allow_html=True)

# ══════════════════ CONSTANTS ══════════════════
MONTH_ORDER = ["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar"]
BLUE_SEQ = [[0,'#f3e5f5'],[0.4,'#9c27b0'],[1,'#6a1b9a']]
CAT_COLORS_LIGHT = ['#7b1fa2','#e91e63','#ff6f00','#1565c0','#2e7d32','#00838f','#f57f17','#6a1b9a','#c62828','#00695c','#4527a0']
GENDER_COLORS = ['#6a1b9a', '#e91e63']

# ══════════════════ HELPERS ══════════════════
def fmt_inr(v):
    if pd.isna(v) or v == 0: return "—"
    v = int(round(float(v)))
    s = str(abs(v)); prefix = "-" if v < 0 else ""
    if len(s) <= 3: return prefix + s
    last3 = s[-3:]; rest = s[:-3]; groups = []
    while len(rest) > 2: groups.append(rest[-2:]); rest = rest[:-2]
    if rest: groups.append(rest)
    groups.reverse()
    return prefix + ','.join(groups) + ',' + last3

def pct(v, dec=2):
    if pd.isna(v) or v == 0: return "—"
    return f"{float(v)*100:.{dec}f}%"

def chart_layout(height=400, title="", xangle=0, show_legend=True):
    return dict(
        paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(245,240,255,0.5)",
        font=dict(color="#1a0030", family="Inter", size=12),
        margin=dict(l=10, r=10, t=55, b=10), height=height,
        title=dict(text=f"<b>{title}</b>", font=dict(color="#1a0030", size=15, family="Plus Jakarta Sans")),
        legend=dict(font=dict(color="#1a0030", size=11), bgcolor="rgba(255,255,255,0.97)",
                    bordercolor="#ddd6fe", borderwidth=1.5, visible=show_legend),
        xaxis=dict(gridcolor="#ede9fe", tickfont=dict(color="#1a0030", size=11, family="Inter"),
                   linecolor="#ddd6fe", tickangle=xangle, showgrid=True),
        yaxis=dict(gridcolor="#ede9fe", tickfont=dict(color="#1a0030", size=11, family="Inter"),
                   linecolor="#ddd6fe", showgrid=True),
    )

def short_store(name):
    return str(name).replace("SSIPL-", "").replace("SSIPL", "").strip(" -")

# ══════════════════ PROCESSING ══════════════════
@st.cache_data(show_spinner=False)
def process(file_bytes):
    df = pd.read_excel(BytesIO(file_bytes), sheet_name="Sheet1")
    df.columns = [str(c).strip() for c in df.columns]
    df = df[df['BRAND'] == 'LEVIS'].copy()
    df['NET SOLD AMOUNT'] = pd.to_numeric(df['NET SOLD AMOUNT'], errors='coerce').fillna(0)
    df['NET SOLD QTY'] = pd.to_numeric(df['NET SOLD QTY'], errors='coerce').fillna(0)
    df['Store Name'] = df['Store Name'].astype(str).str.strip()
    df['CATEGORY'] = df['CATEGORY'].astype(str).str.strip()
    df = df[df['CATEGORY'] != 'NP']
    if 'CLASS' in df.columns:
        df['CLASS'] = df['CLASS'].astype(str).str.strip()
        df = df[df['CLASS'] != 'NP']

    months = [m for m in MONTH_ORDER if m in df['Month'].unique()]
    stores = sorted(df['Store Name'].dropna().unique())
    cats = sorted(df['CATEGORY'].dropna().unique())
    classes = sorted(df['CLASS'].dropna().unique()) if 'CLASS' in df.columns else []

    # SWC — Store-wise
    swc = df.pivot_table(index='Store Name', columns='Month', values='NET SOLD AMOUNT',
                          aggfunc='sum').reindex(columns=months).fillna(0)
    swc['Total Sale'] = swc.sum(axis=1)
    grand = swc['Total Sale'].sum()
    swc['Sale Cont.'] = swc['Total Sale'] / grand if grand else 0
    qty_by_store = df.groupby('Store Name')['NET SOLD QTY'].sum()
    swc['Total Qty'] = qty_by_store.reindex(swc.index).fillna(0)
    if 'Store Status' in df.columns:
        status_map = df.groupby('Store Name')['Store Status'].agg(lambda x: x.dropna().iloc[0] if len(x.dropna()) else "")
        swc['Store Status'] = status_map.reindex(swc.index)

    # CWC — Category-wise
    cwc = df.pivot_table(index='Store Name', columns='CATEGORY', values='NET SOLD AMOUNT',
                          aggfunc='sum').reindex(columns=cats).fillna(0)
    cwc['TOTAL'] = cwc.sum(axis=1)
    gt_cat = cwc.sum()
    cat_cont = (gt_cat / gt_cat['TOTAL']) if gt_cat['TOTAL'] else gt_cat * 0
    cat_month = df.pivot_table(index='CATEGORY', columns='Month', values='NET SOLD AMOUNT',
                                aggfunc='sum').reindex(index=cats, columns=months).fillna(0)

    # CLASS-wise (Jeans/Polos/Tees/etc.)
    class_total = df.groupby('CLASS')['NET SOLD AMOUNT'].sum().sort_values(ascending=False) if classes else pd.Series(dtype=float)
    class_store = df.pivot_table(index='Store Name', columns='CLASS', values='NET SOLD AMOUNT',
                                  aggfunc='sum').fillna(0) if classes else pd.DataFrame()

    # Gender
    gdf = df[df['Gender'].notna()] if 'Gender' in df.columns else df.iloc[0:0]
    gender_total = gdf.groupby('Gender')['NET SOLD AMOUNT'].sum() if len(gdf) else pd.Series(dtype=float)
    gender_month = gdf.pivot_table(index='Gender', columns='Month', values='NET SOLD AMOUNT',
                                    aggfunc='sum').reindex(columns=months).fillna(0) if len(gdf) else pd.DataFrame()
    gender_cat = gdf.pivot_table(index='Gender', columns='CATEGORY', values='NET SOLD AMOUNT',
                                  aggfunc='sum').reindex(columns=cats).fillna(0) if len(gdf) else pd.DataFrame()

    total_qty = df['NET SOLD QTY'].sum()
    total_bills = int(df['Trx'].sum()) if 'Trx' in df.columns else df['TRANSNUM'].nunique()
    avg_bill = grand / total_bills if total_bills else 0

    # ── SIZE CLASSIFICATION — Full Size (numeric waist, e.g. denim) vs Cut Size (S/M/L/XL apparel) ──
    size_type = pd.Series("Unclassified", index=df.index)
    if 'WAIST' in df.columns:
        waist_str = df['WAIST'].astype(str).str.strip()
        is_numeric = pd.to_numeric(waist_str, errors='coerce').notna()
        is_cutsize = waist_str.isin(['XS', 'S', 'M', 'L', 'XL', 'XXL', 'XXXL'])
        size_type = np.where(is_numeric, "Full Size", np.where(is_cutsize, "Cut Size", "Unclassified"))
        size_type = pd.Series(size_type, index=df.index)
    df['_SizeType'] = size_type

    size_total = df[df['_SizeType'] != 'Unclassified'].groupby('_SizeType')['NET SOLD AMOUNT'].sum()
    size_month = df[df['_SizeType'] != 'Unclassified'].pivot_table(
        index='_SizeType', columns='Month', values='NET SOLD AMOUNT', aggfunc='sum'
    ).reindex(columns=months).fillna(0)
    fullsize_waist = df[df['_SizeType'] == 'Full Size'].groupby('WAIST')['NET SOLD AMOUNT'].sum().sort_values(ascending=False)
    cutsize_waist = df[df['_SizeType'] == 'Cut Size'].groupby('WAIST')['NET SOLD AMOUNT'].sum()
    cutsize_order = ['XS', 'S', 'M', 'L', 'XL', 'XXL', 'XXXL']
    cutsize_waist = cutsize_waist.reindex([c for c in cutsize_order if c in cutsize_waist.index]).dropna()

    # Store-level raw sale slice for deep-dive (kept small: only needed columns)
    keep_cols = [c for c in ['Store Name', 'Month', 'CATEGORY', 'CLASS', 'Gender',
                              'NET SOLD AMOUNT', 'NET SOLD QTY', '_SizeType'] if c in df.columns]
    raw_slim = df[keep_cols].copy()

    return dict(swc=swc, grand=grand, months=months, stores=stores, cats=cats, classes=classes,
                cwc=cwc, gt_cat=gt_cat, cat_cont=cat_cont, cat_month=cat_month,
                gender_total=gender_total, gender_month=gender_month, gender_cat=gender_cat,
                total_qty=total_qty, total_bills=total_bills, avg_bill=avg_bill,
                size_total=size_total, size_month=size_month,
                fullsize_waist=fullsize_waist, cutsize_waist=cutsize_waist,
                class_total=class_total, class_store=class_store, raw_slim=raw_slim)


def build_excel(d):
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Store-wise"
    hdr_fill = PatternFill("solid", fgColor="6a1b9a")
    hdr_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
    tot_fill = PatternFill("solid", fgColor="ede9fe")
    tot_font = Font(bold=True, size=10, color="4c1d95", name="Calibri")
    border = Border(*[Side(style='thin', color="e5e7eb")]*4)

    swc = d['swc']
    cols = list(swc.columns)
    ws1.cell(1, 1, "Store Name").fill = hdr_fill; ws1.cell(1, 1).font = hdr_font
    for j, c in enumerate(cols, start=2):
        cell = ws1.cell(1, j, c); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
        cell.alignment = Alignment(horizontal="center")
    for i, (store, row) in enumerate(swc.iterrows(), start=2):
        ws1.cell(i, 1, store).border = border
        for j, c in enumerate(cols, start=2):
            v = row[c]
            if c == "Sale Cont.":
                cell = ws1.cell(i, j, float(v)); cell.number_format = '0.00%'
            elif c == "Store Status":
                cell = ws1.cell(i, j, v)
            else:
                cell = ws1.cell(i, j, round(float(v), 0) if pd.notna(v) else 0); cell.number_format = '#,##0'
            cell.border = border
    ws1.freeze_panes = "B2"
    ws1.column_dimensions['A'].width = 32
    for j in range(2, len(cols)+2): ws1.column_dimensions[get_column_letter(j)].width = 14

    ws2 = wb.create_sheet("Category-wise")
    cwc = d['cwc']
    ccols = list(cwc.columns)
    ws2.cell(1, 1, "Store Name").fill = hdr_fill; ws2.cell(1, 1).font = hdr_font
    for j, c in enumerate(ccols, start=2):
        cell = ws2.cell(1, j, c); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
        cell.alignment = Alignment(horizontal="center")
    for i, (store, row) in enumerate(cwc.iterrows(), start=2):
        ws2.cell(i, 1, store).border = border
        for j, c in enumerate(ccols, start=2):
            cell = ws2.cell(i, j, round(float(row[c]), 0)); cell.number_format = '#,##0'; cell.border = border
    gt_row = len(cwc) + 2
    ws2.cell(gt_row, 1, "Grand Total").fill = tot_fill; ws2.cell(gt_row, 1).font = tot_font
    for j, c in enumerate(ccols, start=2):
        cell = ws2.cell(gt_row, j, round(float(d['gt_cat'][c]), 0))
        cell.fill = tot_fill; cell.font = tot_font; cell.number_format = '#,##0'
    ws2.freeze_panes = "B2"
    ws2.column_dimensions['A'].width = 32
    for j in range(2, len(ccols)+2): ws2.column_dimensions[get_column_letter(j)].width = 16

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ══════════════════ HEADER ══════════════════
st.markdown("""
<div class="hero">
    <div class="hero-badge">LEVI'S ANALYZER</div>
    <div class="hero-divider"></div>
    <div class="hero-title">Levi's Sale Analyzer</div>
    <div class="hero-arrow">▸</div>
    <div style="flex:1; overflow:hidden;">
        <div class="hero-sub-line">Store · Category · Gender · Size · Cut Size</div>
        <div class="hero-sub">Upload RAW sale export · Auto Reports · Interactive Dashboard</div>
    </div>
</div>
""", unsafe_allow_html=True)

uploaded = st.file_uploader("Upload Levi's RAW Sale Data (Excel — Sheet1 raw transaction export)", type=["xlsx", "xls"])

if uploaded is None:
    st.info("👆 Upload the Levi's sale data Excel file to get started. Expected sheet: **Sheet1** with columns like `BRAND`, `Store Name`, `Month`, `CATEGORY`, `Gender`, `NET SOLD AMOUNT`, `NET SOLD QTY`.")
    st.stop()

with st.spinner("Processing Levi's sale data..."):
    d = process(uploaded.getvalue())

swc, grand, months, stores, cats = d['swc'], d['grand'], d['months'], d['stores'], d['cats']
cwc, gt_cat, cat_cont, cat_month = d['cwc'], d['gt_cat'], d['cat_cont'], d['cat_month']
gender_total, gender_month, gender_cat = d['gender_total'], d['gender_month'], d['gender_cat']

# ══════════════════ KPI ROW ══════════════════
k1, k2, k3, k4, k5 = st.columns(5)
for col, lbl, val, sub in [
    (k1, "Total Sale", f"₹{fmt_inr(grand)}", f"{len(months)} months"),
    (k2, "Total Qty Sold", f"{fmt_inr(d['total_qty'])}", "units"),
    (k3, "Total Stores", f"{len(stores)}", "active stores"),
    (k4, "Total Bills", f"{fmt_inr(d['total_bills'])}", "transactions"),
    (k5, "Avg Bill Value", f"₹{fmt_inr(d['avg_bill'])}", "per bill"),
]:
    with col:
        st.markdown(f"""<div class="kpi-card"><div class="kpi-label">{lbl}</div>
            <div class="kpi-value">{val}</div><div class="kpi-sub">{sub}</div></div>""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

tabs = st.tabs(["📈 Overview", "🏪 Store-wise", "📦 Category-wise", "📏 Size Analysis", "🚻 Gender-wise",
                "🔥 Heatmap", "🔍 Store Deep Dive", "📊 Performance", "🤖 AI Strategy Summary"])

# ══════════════════ TAB 1: OVERVIEW ══════════════════
with tabs[0]:
    st.markdown('<div class="section-title">📈 Monthly Sale Trend</div>', unsafe_allow_html=True)
    monthly = swc[months].sum()
    best_idx = int(monthly.values.argmax()); worst_idx = int(monthly.values.argmin())
    bar_colors = ['#9c27b0' if i not in [best_idx, worst_idx] else ('#16a34a' if i == best_idx else '#dc2626') for i in range(len(monthly))]
    fig = go.Figure(go.Bar(
        x=months, y=monthly.values, marker=dict(color=bar_colors, line=dict(width=0)),
        text=[f"₹{fmt_inr(v)}" for v in monthly.values], textposition='outside',
        textfont=dict(size=13, color='#1a0030', family='Inter'),
        hovertemplate='<b>%{x}</b><br>Sale: ₹%{y:,.0f}<extra></extra>',
    ))
    fig.update_layout(**chart_layout(380, "Monthly Sale — All Levi's Stores Combined"),
                       bargap=0.3, yaxis_range=[0, monthly.max() * 1.22],
                       annotations=[
                           dict(x=months[best_idx], y=monthly.values[best_idx] * 1.15, text="🏆 Best", showarrow=False, font=dict(color='#16a34a', size=11)),
                           dict(x=months[worst_idx], y=monthly.values[worst_idx] * 1.15, text="⬇ Low", showarrow=False, font=dict(color='#dc2626', size=11)),
                       ])
    st.plotly_chart(fig, use_container_width=True)

    avg_m = monthly.mean()
    growth = ((monthly.values[-1] - monthly.values[0]) / monthly.values[0] * 100) if monthly.values[0] > 0 else 0
    i1, i2, i3, i4 = st.columns(4)
    for col, lbl, val, sub in [
        (i1, "📈 Best Month", months[best_idx], f"₹{fmt_inr(monthly.values[best_idx])}"),
        (i2, "📉 Lowest Month", months[worst_idx], f"₹{fmt_inr(monthly.values[worst_idx])}"),
        (i3, "📊 Avg Monthly", f"₹{fmt_inr(avg_m)}", "Per month average"),
        (i4, f"🚀 {months[0]}→{months[-1]}", f"{growth:+.1f}%", "Growth trend"),
    ]:
        with col:
            st.markdown(f"""<div style="background:#fff;border-radius:12px;padding:.9rem 1.1rem;
                box-shadow:0 2px 10px rgba(106,27,154,.1);border-left:4px solid #9c27b0">
                <div style="font-size:.58rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#6a1b9a">{lbl}</div>
                <div style="font-size:1.25rem;font-weight:800;color:#1a0030;margin:.2rem 0">{val}</div>
                <div style="font-size:.72rem;color:#607d9b">{sub}</div>
            </div>""", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    ca, cb = st.columns(2)
    with ca:
        st.markdown('<div class="section-title">🏆 Top 10 Stores</div>', unsafe_allow_html=True)
        top10 = swc['Total Sale'].nlargest(10).sort_values()
        fig2 = go.Figure(go.Bar(
            x=top10.values, y=[short_store(s) for s in top10.index], orientation='h',
            marker=dict(color=top10.values, colorscale=BLUE_SEQ, line=dict(width=0)),
            text=[f"₹{fmt_inr(v)}" for v in top10.values], textposition='outside',
            textfont=dict(size=12, color='#1a0030', family='Inter'),
            hovertemplate='<b>%{y}</b><br>₹%{x:,.0f}<extra></extra>',
        ))
        fig2.update_layout(**chart_layout(420, "Top 10 Stores by Sale"), xaxis_range=[0, top10.max() * 1.35])
        st.plotly_chart(fig2, use_container_width=True)
    with cb:
        st.markdown('<div class="section-title">📉 Bottom 10 Stores</div>', unsafe_allow_html=True)
        bot10 = swc['Total Sale'].nsmallest(10).sort_values()
        fig4 = go.Figure(go.Bar(
            x=bot10.values, y=[short_store(s) for s in bot10.index], orientation='h',
            marker=dict(color='#dc2626', line=dict(width=0)),
            text=[f"₹{fmt_inr(v)}" for v in bot10.values], textposition='outside',
            textfont=dict(size=12, color='#1a0030', family='Inter'),
            hovertemplate='<b>%{y}</b><br>₹%{x:,.0f}<extra></extra>',
        ))
        fig4.update_layout(**chart_layout(420, "Bottom 10 Stores by Sale"), xaxis_range=[0, bot10.max() * 1.45])
        st.plotly_chart(fig4, use_container_width=True)

    st.markdown('<div class="section-title" style="margin-top:1rem">🏷️ Category Mix</div>', unsafe_allow_html=True)
    fig3 = go.Figure(go.Pie(
        labels=cats, values=gt_cat[cats].values, hole=0.52,
        marker=dict(colors=CAT_COLORS_LIGHT[:len(cats)], line=dict(color='#ffffff', width=3)),
        textinfo='label+percent', textfont=dict(size=12, color='#1a0030', family='Inter'),
        insidetextfont=dict(size=11, color='#ffffff'),
        hovertemplate='<b>%{label}</b><br>₹%{value:,.0f}<br>%{percent}<extra></extra>',
    ))
    fig3.update_layout(**chart_layout(420, "Category-wise Sale Contribution"),
                        annotations=[dict(text=f"<b>₹{fmt_inr(grand)}</b>", x=0.5, y=0.5,
                                           font=dict(size=13, color='#1a0030', family='Plus Jakarta Sans'), showarrow=False)])
    st.plotly_chart(fig3, use_container_width=True)

# ══════════════════ TAB 2: STORE-WISE ══════════════════
with tabs[1]:
    st.markdown('<div class="section-title">🏪 Store-wise Monthly Sale</div>', unsafe_allow_html=True)
    top5_default = swc['Total Sale'].nlargest(5).index.tolist()
    sel_stores = st.multiselect("Select Stores", stores, default=top5_default, key="store_multiselect_levis")
    if sel_stores:
        fig_sw = go.Figure()
        for i, sn in enumerate(sel_stores):
            fig_sw.add_trace(go.Bar(x=months, y=swc.loc[sn, months].values, name=short_store(sn),
                                     marker_color=CAT_COLORS_LIGHT[i % len(CAT_COLORS_LIGHT)],
                                     hovertemplate=f'<b>{short_store(sn)}</b><br>%{{x}}: ₹%{{y:,.0f}}<extra></extra>'))
        fig_sw.update_layout(**chart_layout(430, "Store-wise Monthly Sale"), barmode='group', bargap=0.12)
        st.plotly_chart(fig_sw, use_container_width=True)

    st.markdown('<div class="section-title" style="margin-top:1rem">📋 SWC Table</div>', unsafe_allow_html=True)
    disp = swc.copy()
    disp_fmt = disp.copy()
    for m in months: disp_fmt[m] = disp[m].apply(fmt_inr)
    disp_fmt['Total Sale'] = disp['Total Sale'].apply(fmt_inr)
    disp_fmt['Sale Cont.'] = disp['Sale Cont.'].apply(pct)
    disp_fmt['Total Qty'] = disp['Total Qty'].apply(lambda v: fmt_inr(v))
    st.dataframe(disp_fmt.sort_values('Total Sale', ascending=False, key=lambda x: disp['Total Sale']), use_container_width=True, height=420)

    excel_buf = build_excel(d)
    st.download_button("⬇ Download Full Report (Excel)", data=excel_buf,
                        file_name="Levis_Sale_Analyzer_Report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══════════════════ TAB 3: CATEGORY-WISE ══════════════════
with tabs[2]:
    st.markdown('<div class="section-title">📦 Category-wise Sale Contribution (CWC)</div>', unsafe_allow_html=True)
    cwc_fmt = cwc.copy()
    for c in cwc.columns: cwc_fmt[c] = cwc[c].apply(fmt_inr)
    st.dataframe(cwc_fmt, use_container_width=True, height=380)

    ca, cb = st.columns(2)
    with ca:
        st.markdown('<div class="section-title">📊 Category Monthly Trend</div>', unsafe_allow_html=True)
        fig5 = go.Figure()
        for i, c in enumerate(cats):
            fig5.add_trace(go.Scatter(x=months, y=cat_month.loc[c].values, mode='lines+markers',
                                       name=c, line=dict(color=CAT_COLORS_LIGHT[i % len(CAT_COLORS_LIGHT)], width=2.5)))
        fig5.update_layout(**chart_layout(400, "Category Sale by Month"))
        st.plotly_chart(fig5, use_container_width=True)
    with cb:
        st.markdown('<div class="section-title">🏷️ Category Contribution %</div>', unsafe_allow_html=True)
        cat_sorted = gt_cat[cats].sort_values()
        fig6 = go.Figure(go.Bar(
            x=cat_sorted.values, y=cat_sorted.index, orientation='h',
            marker=dict(color=CAT_COLORS_LIGHT[:len(cats)]),
            text=[pct(v/grand) for v in cat_sorted.values], textposition='outside',
            textfont=dict(size=12, color='#1a0030', family='Inter'),
        ))
        fig6.update_layout(**chart_layout(400, "Category Contribution"), xaxis_range=[0, cat_sorted.max()*1.35])
        st.plotly_chart(fig6, use_container_width=True)

    class_total = d['class_total']
    if len(class_total):
        st.markdown('<div class="section-title" style="margin-top:1rem">👖 Class-wise Sale (Jeans · Polos · Tees · Shorts...)</div>', unsafe_allow_html=True)
        top_classes = class_total.nlargest(15).sort_values()
        fig6b = go.Figure(go.Bar(
            x=top_classes.values, y=top_classes.index, orientation='h',
            marker=dict(color=top_classes.values, colorscale=BLUE_SEQ),
            text=[f"₹{fmt_inr(v)}" for v in top_classes.values], textposition='outside',
            textfont=dict(size=11, color='#1a0030', family='Inter'),
        ))
        fig6b.update_layout(**chart_layout(460, "Top Classes by Sale"), xaxis_range=[0, top_classes.max()*1.35])
        st.plotly_chart(fig6b, use_container_width=True)

# ══════════════════ TAB 4: SIZE ANALYSIS ══════════════════
with tabs[3]:
    size_total, size_month = d['size_total'], d['size_month']
    fullsize_waist, cutsize_waist = d['fullsize_waist'], d['cutsize_waist']
    if len(size_total) == 0:
        st.info("Size (WAIST) data not available in this file.")
    else:
        st.markdown('<div class="section-title">📏 Full Size vs Cut Size Sale Split</div>', unsafe_allow_html=True)
        ca, cb = st.columns(2)
        with ca:
            fig10 = go.Figure(go.Pie(
                labels=size_total.index, values=size_total.values, hole=0.5,
                marker=dict(colors=['#6a1b9a', '#f57f17'], line=dict(color='#ffffff', width=3)),
                textinfo='label+percent', textfont=dict(size=14, color='#1a0030', family='Inter'),
                insidetextfont=dict(size=13, color='#ffffff'),
                hovertemplate='<b>%{label}</b><br>₹%{value:,.0f}<br>%{percent}<extra></extra>',
            ))
            fig10.update_layout(**chart_layout(380, "Full Size vs Cut Size — Total Sale"))
            st.plotly_chart(fig10, use_container_width=True)
        with cb:
            st.markdown('<div class="section-title">📈 Monthly Trend</div>', unsafe_allow_html=True)
            fig11 = go.Figure()
            for i, s in enumerate(size_month.index):
                fig11.add_trace(go.Bar(x=months, y=size_month.loc[s].values, name=s,
                                        marker_color=['#6a1b9a', '#f57f17'][i % 2]))
            fig11.update_layout(**chart_layout(380, "Size Type — Monthly Sale"), barmode='group')
            st.plotly_chart(fig11, use_container_width=True)

        st.markdown('<div class="section-title" style="margin-top:1rem">📐 Full Size — Waist-wise Sale (Denim/Bottoms)</div>', unsafe_allow_html=True)
        if len(fullsize_waist):
            fig12 = go.Figure(go.Bar(
                x=[str(w) for w in fullsize_waist.index], y=fullsize_waist.values,
                marker=dict(color=fullsize_waist.values, colorscale=BLUE_SEQ),
                text=[f"₹{fmt_inr(v)}" for v in fullsize_waist.values], textposition='outside',
                textfont=dict(size=11, color='#1a0030', family='Inter'),
            ))
            fig12.update_layout(**chart_layout(360, "Sale by Waist Size"))
            st.plotly_chart(fig12, use_container_width=True)
        else:
            st.info("No Full Size (denim waist) data found.")

        st.markdown('<div class="section-title" style="margin-top:1rem">👕 Cut Size — S/M/L/XL-wise Sale (Apparel)</div>', unsafe_allow_html=True)
        if len(cutsize_waist):
            fig13 = go.Figure(go.Bar(
                x=cutsize_waist.index, y=cutsize_waist.values,
                marker=dict(color='#e91e63'),
                text=[f"₹{fmt_inr(v)}" for v in cutsize_waist.values], textposition='outside',
                textfont=dict(size=12, color='#1a0030', family='Inter'),
            ))
            fig13.update_layout(**chart_layout(360, "Sale by Cut Size"))
            st.plotly_chart(fig13, use_container_width=True)
        else:
            st.info("No Cut Size (S/M/L/XL) data found.")

# ══════════════════ TAB 5: GENDER-WISE ══════════════════
with tabs[4]:
    if len(gender_total) == 0:
        st.info("Gender data not available in this file.")
    else:
        st.markdown('<div class="section-title">🚻 Gender-wise Sale Split</div>', unsafe_allow_html=True)
        ca, cb = st.columns(2)
        with ca:
            fig7 = go.Figure(go.Pie(
                labels=gender_total.index, values=gender_total.values, hole=0.5,
                marker=dict(colors=GENDER_COLORS, line=dict(color='#ffffff', width=3)),
                textinfo='label+percent', textfont=dict(size=14, color='#1a0030', family='Inter'),
                insidetextfont=dict(size=13, color='#ffffff'),
                hovertemplate='<b>%{label}</b><br>₹%{value:,.0f}<br>%{percent}<extra></extra>',
            ))
            fig7.update_layout(**chart_layout(380, "Mens vs Womens — Total Sale"))
            st.plotly_chart(fig7, use_container_width=True)
        with cb:
            st.markdown('<div class="section-title">📈 Gender Monthly Trend</div>', unsafe_allow_html=True)
            fig8 = go.Figure()
            for i, g in enumerate(gender_month.index):
                fig8.add_trace(go.Bar(x=months, y=gender_month.loc[g].values, name=g,
                                       marker_color=GENDER_COLORS[i % len(GENDER_COLORS)]))
            fig8.update_layout(**chart_layout(380, "Gender-wise Monthly Sale"), barmode='group')
            st.plotly_chart(fig8, use_container_width=True)

        st.markdown('<div class="section-title" style="margin-top:1rem">📦 Gender × Category Breakdown</div>', unsafe_allow_html=True)
        fig9 = go.Figure()
        for i, g in enumerate(gender_cat.index):
            fig9.add_trace(go.Bar(x=gender_cat.columns, y=gender_cat.loc[g].values, name=g,
                                   marker_color=GENDER_COLORS[i % len(GENDER_COLORS)]))
        fig9.update_layout(**chart_layout(380, "Category Sale by Gender", xangle=-20), barmode='group')
        st.plotly_chart(fig9, use_container_width=True)

# ══════════════════ TAB 6: HEATMAP ══════════════════
with tabs[5]:
    st.markdown('<div class="section-title">🔥 Sale Heatmap — Store × Category</div>', unsafe_allow_html=True)
    hm = d['cwc'][cats].copy()
    hm_nan = hm.replace(0, np.nan)
    fig_hm = go.Figure(go.Heatmap(
        z=hm_nan.values.tolist(), x=hm_nan.columns.tolist(), y=hm_nan.index.tolist(),
        colorscale=[[0, '#fdf8ff'], [0.2, '#e9d8f8'], [0.5, '#c084fc'], [0.75, '#9333ea'], [1, '#581c87']],
        text=[[f"₹{fmt_inr(v)}" if pd.notna(v) else "—" for v in row] for row in hm_nan.values.tolist()],
        texttemplate="%{text}", textfont=dict(size=9, color='#1a0030'),
        hoverongaps=False, colorbar=dict(title="Sale", tickfont=dict(color='#1a0030')),
    ))
    fig_hm.update_layout(
        paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(245,240,255,0.5)",
        font=dict(color="#1a0030", family="Inter", size=11), height=max(500, len(stores) * 22),
        margin=dict(l=200, r=20, t=55, b=80),
        title=dict(text="<b>Sale Heatmap: Store × Category</b>", font=dict(color='#1a0030', size=14, family='Plus Jakarta Sans')),
        xaxis=dict(tickangle=-30, tickfont=dict(size=10, color='#1a0030')),
        yaxis=dict(tickfont=dict(size=9, color='#1a0030'), autorange='reversed'),
    )
    st.plotly_chart(fig_hm, use_container_width=True)

    if hm.values.max() > 0:
        mi = np.unravel_index(hm.values.argmax(), hm.shape)
        rt = hm.sum(axis=1).sort_values(); ct = hm.sum(axis=0).sort_values(ascending=False)
        st.markdown(f"""<div style="background:#f8faff;border:1.5px solid #c7d7f9;border-radius:12px;padding:1rem 1.2rem;margin-top:.8rem">
          <div style="font-size:.6rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#1e40af;margin-bottom:.7rem">🔥 HEATMAP — KEY INSIGHTS</div>
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:.7rem">
            <div style="background:#f5f3ff;border-radius:8px;padding:.7rem;border-left:4px solid #7c3aed">
              <div style="font-size:.65rem;font-weight:700;color:#4c1d95;margin-bottom:.3rem">🏆 HIGHEST COMBINATION</div>
              <div style="font-size:.85rem;font-weight:800;color:#1a0030"><b>{short_store(hm.index[mi[0]])}</b> → {hm.columns[mi[1]]}</div>
              <div style="font-size:.78rem;color:#4c1d95">₹{fmt_inr(hm.values[mi])}</div>
            </div>
            <div style="background:#f0fdf4;border-radius:8px;padding:.7rem;border-left:4px solid #16a34a">
              <div style="font-size:.65rem;font-weight:700;color:#166534;margin-bottom:.3rem">📦 TOP 3 CATEGORIES</div>
              <div style="font-size:.78rem;color:#1a0030">{"<br>".join([f"<b>{c}</b> — ₹{fmt_inr(v)}" for c, v in ct.head(3).items()])}</div>
            </div>
            <div style="background:#eff6ff;border-radius:8px;padding:.7rem;border-left:4px solid #1d4ed8">
              <div style="font-size:.65rem;font-weight:700;color:#1e40af;margin-bottom:.3rem">🏆 BEST STORE</div>
              <div style="font-size:.78rem;color:#1a0030"><b>{short_store(rt.index[-1])}</b> — ₹{fmt_inr(rt.values[-1])}</div>
            </div>
            <div style="background:#fef2f2;border-radius:8px;padding:.7rem;border-left:4px solid #dc2626">
              <div style="font-size:.65rem;font-weight:700;color:#991b1b;margin-bottom:.3rem">⚠️ WEAKEST STORE</div>
              <div style="font-size:.78rem;color:#1a0030"><b>{short_store(rt.index[0])}</b> — ₹{fmt_inr(rt.values[0])}</div>
            </div>
          </div>
        </div>""", unsafe_allow_html=True)

# ══════════════════ TAB 7: STORE DEEP DIVE ══════════════════
with tabs[6]:
    st.markdown('<div class="section-title">🔍 Store Deep Dive</div>', unsafe_allow_html=True)
    raw_slim = d['raw_slim']
    dd_store = st.selectbox("Select Store", stores, key="dd_store_levis")
    if dd_store:
        ss_df = raw_slim[raw_slim['Store Name'] == dd_store]
        ts = ss_df['NET SOLD AMOUNT'].sum()
        tq = ss_df['NET SOLD QTY'].sum()
        rank = int(swc['Total Sale'].rank(ascending=False)[dd_store])
        cont = ts / grand if grand else 0

        m1, m2, m3, m4 = st.columns(4)
        for col, lbl, val, sub in [
            (m1, "Total Sale", f"₹{fmt_inr(ts)}", f"{len(months)} months"),
            (m2, "Total Qty", f"{fmt_inr(tq)}", "units sold"),
            (m3, "Contribution", pct(cont, 4), "of total sale"),
            (m4, "Store Rank", f"#{rank}", f"out of {len(stores)} stores"),
        ]:
            with col:
                st.markdown(f"""<div class="kpi-card"><div class="kpi-label">{lbl}</div>
                    <div class="kpi-value">{val}</div><div class="kpi-sub">{sub}</div></div>""", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

        mon_opts = ["All Months"] + months
        sel_mon = st.radio("📅 Select Month", mon_opts, horizontal=True, key=f"dd_mon_levis_{dd_store}")
        ss_filter = ss_df if sel_mon == "All Months" else ss_df[ss_df['Month'] == sel_mon]
        title_suffix = " — All Months" if sel_mon == "All Months" else f" — {sel_mon}"

        d1, d2 = st.columns([3, 2])
        with d1:
            mm = ss_df.groupby('Month')['NET SOLD AMOUNT'].sum().reindex(months).fillna(0)
            bar_clrs = ['#9c27b0'] * len(months)
            if sel_mon != "All Months":
                bar_clrs = ['#f3e5f5'] * len(months)
                bar_clrs[months.index(sel_mon)] = '#6a1b9a'
            fig_dm = go.Figure(go.Bar(x=months, y=mm.values,
                marker=dict(color=bar_clrs, line=dict(width=0)),
                text=[f"₹{fmt_inr(v)}" if v > 0 else "" for v in mm.values],
                textposition='outside', textfont=dict(size=10, color='#1a0030')))
            fig_dm.update_layout(**chart_layout(280, f"{short_store(dd_store)} — Monthly Sale"), bargap=0.3)
            st.plotly_chart(fig_dm, use_container_width=True)
        with d2:
            cd = ss_filter.groupby('CATEGORY')['NET SOLD AMOUNT'].sum(); cd = cd[cd > 0]
            if len(cd):
                fig_dp = go.Figure(go.Pie(labels=cd.index.tolist(), values=cd.values.tolist(), hole=0.48,
                    marker=dict(colors=CAT_COLORS_LIGHT[:len(cd)], line=dict(color='#fff', width=2)),
                    textinfo='label+percent', textfont=dict(size=11, color='#1a0030'),
                    insidetextfont=dict(size=10, color='#fff')))
                fig_dp.update_layout(**chart_layout(280, f"Category Mix{title_suffix}"))
                st.plotly_chart(fig_dp, use_container_width=True)

        d3, d4 = st.columns(2)
        with d3:
            if 'CLASS' in ss_filter.columns:
                bd = ss_filter.groupby('CLASS')['NET SOLD AMOUNT'].sum().sort_values(ascending=False); bd = bd[bd > 0].head(8)
                fig_db = go.Figure(go.Bar(x=bd.index.tolist(), y=bd.values.tolist(),
                    marker=dict(color=CAT_COLORS_LIGHT[:len(bd)]),
                    text=[f"₹{fmt_inr(v)}" for v in bd.values], textposition='outside'))
                fig_db.update_layout(**chart_layout(280, f"Class Mix{title_suffix}", xangle=-30), bargap=0.3)
                st.plotly_chart(fig_db, use_container_width=True)
        with d4:
            gd = ss_filter.groupby('Gender')['NET SOLD AMOUNT'].sum().sort_values(ascending=False) if 'Gender' in ss_filter.columns else pd.Series(dtype=float)
            gd = gd[gd > 0]
            if len(gd):
                fig_dg = go.Figure(go.Pie(labels=gd.index.tolist(), values=gd.values.tolist(), hole=0.48,
                    marker=dict(colors=GENDER_COLORS, line=dict(color='#fff', width=2)),
                    textinfo='label+percent', textfont=dict(size=12, color='#1a0030'),
                    insidetextfont=dict(size=10, color='#fff')))
                fig_dg.update_layout(**chart_layout(280, f"Gender Mix{title_suffix}"))
                st.plotly_chart(fig_dg, use_container_width=True)

        if sel_mon != "All Months":
            st.markdown(f'<div class="section-title" style="margin-top:1rem">📋 Full Detail — {sel_mon}</div>', unsafe_allow_html=True)
            detail_cols = [c for c in ['CATEGORY', 'CLASS', 'Gender'] if c in ss_filter.columns]
            detail = ss_filter.groupby(detail_cols)['NET SOLD AMOUNT'].sum().reset_index()
            detail = detail[detail['NET SOLD AMOUNT'] > 0].sort_values('NET SOLD AMOUNT', ascending=False)
            detail['NET SOLD AMOUNT'] = detail['NET SOLD AMOUNT'].apply(fmt_inr)
            detail = detail.rename(columns={'NET SOLD AMOUNT': 'Sale'})
            st.dataframe(detail, use_container_width=True, hide_index=True)

# ══════════════════ TAB 8: PERFORMANCE ══════════════════
with tabs[7]:
    st.markdown('<div class="section-title">📊 Month-on-Month Growth</div>', unsafe_allow_html=True)
    mall = swc[months].sum()
    mom = mall.pct_change() * 100
    mp = [float(v) for v in mom.values[1:]]
    if mp:
        fig_mom = go.Figure(go.Bar(x=months[1:], y=mp,
            marker=dict(color=['#16a34a' if v >= 0 else '#dc2626' for v in mp]),
            text=[f"{v:+.1f}%" for v in mp], textposition='outside', textfont=dict(size=12, color='#1a0030')))
        fig_mom.update_layout(**chart_layout(340, "MoM Sale Growth (%) — All Stores"), bargap=0.3)
        fig_mom.update_layout(yaxis=dict(gridcolor='#ede9fe', zeroline=True, zerolinecolor='#9c27b0', zerolinewidth=2))
        st.plotly_chart(fig_mom, use_container_width=True)
    else:
        st.info("Need at least 2 months of data for MoM growth.")

    p1, p2 = st.columns(2)
    with p1:
        st.markdown('<div class="section-title">🏆 Top 5 Stores</div>', unsafe_allow_html=True)
        t5 = swc['Total Sale'].nlargest(5).sort_values()
        fig_t5 = go.Figure(go.Bar(x=t5.values, y=[short_store(s) for s in t5.index], orientation='h',
            marker=dict(color='#16a34a'), text=[f"₹{fmt_inr(v)}" for v in t5.values], textposition='outside'))
        fig_t5.update_layout(**chart_layout(300, "Top 5 Stores"), xaxis_range=[0, t5.max() * 1.45])
        st.plotly_chart(fig_t5, use_container_width=True)
    with p2:
        st.markdown('<div class="section-title">🔴 Bottom 5 Stores</div>', unsafe_allow_html=True)
        b5 = swc['Total Sale'].nsmallest(5).sort_values(ascending=False)
        fig_b5 = go.Figure(go.Bar(x=b5.values, y=[short_store(s) for s in b5.index], orientation='h',
            marker=dict(color='#dc2626'), text=[f"₹{fmt_inr(v)}" for v in b5.values], textposition='outside'))
        fig_b5.update_layout(**chart_layout(300, "Bottom 5 Stores"), xaxis_range=[0, b5.max() * 1.55])
        st.plotly_chart(fig_b5, use_container_width=True)

    st.markdown('<div class="section-title" style="margin-top:1rem">❌ Zero Sale Months — Store-wise</div>', unsafe_allow_html=True)
    zr = []
    for sn in swc.index:
        row = swc.loc[sn, months]
        zm = [months[i] for i, v in enumerate(row.values) if v == 0]
        sm = [months[i] for i, v in enumerate(row.values) if v > 0]
        if zm:
            last = float(row.values[-1]); prev = float(row.values[-2]) if len(row.values) > 1 else 0
            g = f"{'▲' if last >= prev else '▼'} {abs((last - prev) / prev * 100):.1f}%" if prev > 0 else "N/A"
            zr.append({'Store': short_store(sn), 'Sale Months': ', '.join(sm), 'Zero Months': ', '.join(zm),
                       'Zero Count': len(zm), 'Growth': g, 'Total Sale': f"₹{fmt_inr(swc.loc[sn, 'Total Sale'])}"})
    if zr:
        st.dataframe(pd.DataFrame(zr).sort_values('Zero Count', ascending=False), use_container_width=True, hide_index=True)
    else:
        st.success("✅ No zero-sale months — every store sold every month!")

# ══════════════════ TAB 9: AI STRATEGY SUMMARY ══════════════════
with tabs[8]:
    st.markdown('<div class="section-title">🤖 AI Strategy Summary</div>', unsafe_allow_html=True)
    st.markdown("""<div style="background:linear-gradient(135deg,#3a0068,#6a1b9a);border-radius:12px;
        padding:1rem 1.4rem;margin-bottom:1rem;color:#fff">
        <div style="font-size:.65rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;
            color:rgba(255,255,255,.7);margin-bottom:.3rem">HOW IT WORKS</div>
        <div style="font-size:.85rem">AI analyses your complete Levi's sales data — stores, categories,
        gender mix, monthly trend — and generates a business strategy with action plan.</div>
    </div>""", unsafe_allow_html=True)

    top3 = swc['Total Sale'].nlargest(3)
    bot3 = swc['Total Sale'].nsmallest(3)
    cat_sale_sorted = gt_cat[cats].sort_values(ascending=False)
    monthly_vals = swc[months].sum()
    mom_last = ((monthly_vals.values[-1] - monthly_vals.values[-2]) / monthly_vals.values[-2] * 100) if len(monthly_vals) > 1 and monthly_vals.values[-2] > 0 else 0
    gender_line = ", ".join([f"{g}: ₹{fmt_inr(v)} ({v/grand*100:.1f}%)" for g, v in gender_total.items()]) if len(gender_total) else "Not available"

    data_prompt = f"""You are a senior retail business consultant analyzing sales data for Levi's stores (part of SSIPL Group, India).

BUSINESS DATA SUMMARY:
- Total Sale: ₹{fmt_inr(int(grand))} across {len(stores)} stores, {len(cats)} categories, months: {', '.join(months)}
- Total Qty Sold: {fmt_inr(d['total_qty'])} units, Total Bills: {fmt_inr(d['total_bills'])}, Avg Bill Value: ₹{fmt_inr(d['avg_bill'])}
- Last Month Growth: {mom_last:+.1f}%
- Gender Split: {gender_line}

TOP 3 STORES: {', '.join([f"{short_store(s)}=₹{fmt_inr(v)}" for s, v in top3.items()])}
BOTTOM 3 STORES: {', '.join([f"{short_store(s)}=₹{fmt_inr(v)}" for s, v in bot3.items()])}
CATEGORY SALE (high to low): {', '.join([f"{c}=₹{fmt_inr(v)}" for c, v in cat_sale_sorted.items()])}

Based on this data, provide a structured business strategy report in English with these exact sections:

1. EXECUTIVE SUMMARY (2-3 sentences — overall business health)
2. KEY STRENGTHS (3 bullet points — what is working well)
3. CRITICAL ISSUES (3 bullet points — biggest problems)
4. HOW TO INCREASE SALE — STRATEGIES (5 specific, actionable strategies with expected impact)
5. GENDER & CATEGORY FOCUS (which segments to double down on)
6. IMMEDIATE PRIORITIES (3 actions for THIS WEEK, 3 for THIS MONTH, 3 for NEXT QUARTER)

Be specific with store names and numbers. Be direct and business-focused. No fluff."""

    if 'ai_summary_levis' not in st.session_state: st.session_state.ai_summary_levis = None
    if 'ai_loading_levis' not in st.session_state: st.session_state.ai_loading_levis = False

    col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
    with col_btn2:
        if st.button("🤖  Generate AI Strategy Summary", use_container_width=True):
            st.session_state.ai_loading_levis = True
            st.session_state.ai_summary_levis = None

    if st.session_state.ai_loading_levis and st.session_state.ai_summary_levis is None:
        with st.spinner("🤖 AI is analysing your data... please wait..."):
            try:
                import requests
                response = requests.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"Content-Type": "application/json"},
                    json={"model": "claude-sonnet-4-20250514", "max_tokens": 1500,
                          "messages": [{"role": "user", "content": data_prompt}]},
                    timeout=60,
                )
                if response.status_code == 200:
                    result = response.json()
                    st.session_state.ai_summary_levis = result['content'][0]['text']
                else:
                    st.session_state.ai_summary_levis = f"API Error: {response.status_code} — {response.text[:200]}"
            except Exception as e:
                st.session_state.ai_summary_levis = f"Error: {str(e)}"
            st.session_state.ai_loading_levis = False
        st.rerun()

    if st.session_state.ai_summary_levis:
        st.markdown(f"""<div style="background:#fff;border-radius:12px;padding:1.2rem 1.4rem;
            box-shadow:0 2px 10px rgba(106,27,154,.1);white-space:pre-wrap;font-size:.88rem;color:#1a0030;">
            {st.session_state.ai_summary_levis}</div>""", unsafe_allow_html=True)
